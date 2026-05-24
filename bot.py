import asyncio
import os
from datetime import datetime

from aiogram import Bot, Dispatcher
from aiogram.filters import Command
from aiogram.types import Message, FSInputFile
from dotenv import load_dotenv
from openpyxl import Workbook

from db import (
    init_db,
    money,
    add_or_update_partner,
    get_partner_by_telegram_id,
    get_partners,
    create_deal,
    get_deal,
    get_deals,
    update_deal_status,
    add_deal_transaction,
    get_deal_transactions,
    get_deal_totals,
    close_deal_and_calculate_payments,
    get_unpaid_payments,
    mark_payment_paid,
    get_global_capital,
    get_report,
    get_history,
    get_export_data,
    clear_test_data,
    reset_all_data,
    delete_deal_by_id
)


load_dotenv()

BOT_TOKEN = os.getenv("BOT_TOKEN")

if BOT_TOKEN is None:
    raise ValueError("Не найден BOT_TOKEN в .env или в переменных окружения")


bot = Bot(token=BOT_TOKEN)
dp = Dispatcher()


ALLOWED_STATUSES = {
    "куплен": "куплен",
    "проверка": "на проверке",
    "ремонт": "в ремонте",
    "готов": "готов к продаже",
    "выставлен": "выставлен",
    "бронь": "забронирован",
    "продан": "продан",
    "закрыт": "закрыт"
}


def user_data(message: Message):
    # Получаем Telegram-данные человека, который написал команду
    user = message.from_user

    return {
        "telegram_id": user.id,
        "username": user.username,
        "first_name": user.first_name,
        "last_name": user.last_name
    }


def user_view(display_name: str, username: str | None):
    # Красиво отображаем партнёра
    if username:
        return f"{display_name} (@{username})"

    return display_name


def require_partner(message: Message):
    # Проверяем, зарегистрирован ли пользователь как партнёр
    data = user_data(message)
    partner = get_partner_by_telegram_id(data["telegram_id"])

    if partner is None:
        return None

    return partner


def require_admin(message: Message):
    # Проверяем, является ли пользователь админом
    partner = require_partner(message)

    if partner is None:
        return None

    is_admin = partner[6]

    if is_admin != 1:
        return None

    return partner


def deal_status_icon(status: str):
    icons = {
        "куплен": "🛒",
        "на проверке": "🔎",
        "в ремонте": "🛠",
        "готов к продаже": "✅",
        "выставлен": "📢",
        "забронирован": "🤝",
        "продан": "💰",
        "закрыт": "🏁"
    }

    return icons.get(status, "📦")


def type_icon(transaction_type: str):
    if transaction_type == "expense":
        return "🔴 Расход"

    if transaction_type == "income":
        return "🟢 Доход"

    return "⚪ Операция"


@dp.message(Command("bb_start"))
async def start_handler(message: Message):
    await message.answer(
        "📱 Бот-бухгалтер для перепродажи техники\n\n"
        "Я помогаю вести сделки, считать прибыль, долги между партнёрами, "
        "капитал в товаре и делать отчёты.\n\n"
        "🚀 Быстрый старт:\n"
        "1. Каждый партнёр пишет /bb_join\n"
        "2. Создаёте сделку: /bb_new_deal iPhone 12 128GB\n"
        "3. Добавляете расходы: /bb_expense 1 28000 покупка\n"
        "4. Добавляете продажу: /bb_income 1 38000 продажа\n"
        "5. Закрываете сделку: /bb_close 1\n\n"
        "📘 Подробная инструкция: /bb_manual\n"
        "📋 Список команд: /bb_commands"
    )


@dp.message(Command("bb_commands"))
async def commands_handler(message: Message):
    await message.answer(
        "📋 Команды бота-бухгалтера\n\n"
        "👥 Партнёры:\n"
        "/bb_join — зарегистрироваться\n"
        "/bb_me — мой профиль\n"
        "/bb_partners — список партнёров\n\n"
        "📦 Сделки:\n"
        "/bb_new_deal название — создать сделку\n"
        "/bb_deals — список сделок\n"
        "/bb_deal номер — карточка сделки\n"
        "/bb_status номер статус — изменить статус\n\n"
        "💸 Деньги:\n"
        "/bb_expense номер сумма комментарий — расход\n"
        "/bb_income номер сумма комментарий — доход\n"
        "/bb_close номер — закрыть сделку\n"
        "/bb_debts — долги\n"
        "/bb_pay номер — отметить долг оплаченным\n\n"
        "📊 Отчёты:\n"
        "/bb_stock — техника в работе\n"
        "/bb_capital — капитал\n"
        "/bb_day — отчёт за день\n"
        "/bb_week — отчёт за неделю\n"
        "/bb_month — отчёт за месяц\n"
        "/bb_export — Excel-файл\n"
        "/bb_history — история операций\n\n"
        "🧹 Очистка:\n"
        "/bb_clear — удалить сделки, операции и долги\n"
        "/bb_delete_deal 1 — удалить конкретную сделку\n"
        "/bb_reset — полный сброс базы\n\n"
        "📘 Инструкция:\n"
        "/bb_manual"
    )


@dp.message(Command("bb_manual"))
async def manual_handler(message: Message):
    await message.answer(
        "📘 Подробная инструкция\n\n"
        "1️⃣ Регистрация партнёров\n\n"
        "Каждый человек из команды должен написать:\n"
        "/bb_join\n\n"
        "Бот запомнит Telegram-аккаунт. После этого не надо писать имя вручную — "
        "бот сам понимает, кто добавил расход или доход.\n\n"
        "2️⃣ Создание сделки\n\n"
        "Одна единица техники = одна сделка.\n\n"
        "Пример:\n"
        "/bb_new_deal iPhone 12 128GB Black\n\n"
        "Бот создаст сделку и выдаст номер, например #1.\n\n"
        "3️⃣ Добавление расходов\n\n"
        "Расходы — покупка, ремонт, доставка, аксессуары, комиссии.\n\n"
        "Примеры:\n"
        "/bb_expense 1 28000 покупка\n"
        "/bb_expense 1 2000 замена АКБ\n"
        "/bb_expense 1 500 доставка\n\n"
        "Расход записывается на того, кто отправил команду.\n\n"
        "4️⃣ Добавление дохода\n\n"
        "Когда устройство продали:\n"
        "/bb_income 1 38000 продажа\n\n"
        "Доход записывается на того, кто получил деньги от покупателя.\n\n"
        "5️⃣ Статусы\n\n"
        "Менять статус:\n"
        "/bb_status 1 ремонт\n"
        "/bb_status 1 выставлен\n"
        "/bb_status 1 продан\n\n"
        "Доступные статусы:\n"
        "куплен, проверка, ремонт, готов, выставлен, бронь, продан, закрыт\n\n"
        "6️⃣ Закрытие сделки\n\n"
        "Когда все расходы и доходы внесены:\n"
        "/bb_close 1\n\n"
        "Бот посчитает расходы, доходы, чистую прибыль, долю каждого партнёра "
        "и кто кому должен.\n\n"
        "7️⃣ Долги\n\n"
        "Посмотреть долги:\n"
        "/bb_debts\n\n"
        "Отметить долг как оплаченный:\n"
        "/bb_pay 1\n\n"
        "8️⃣ Отчёты\n\n"
        "/bb_stock — техника в работе\n"
        "/bb_capital — деньги в товаре и общая прибыль\n"
        "/bb_day — отчёт за день\n"
        "/bb_week — отчёт за неделю\n"
        "/bb_month — отчёт за месяц\n"
        "/bb_export — выгрузка Excel\n\n"
        "9️⃣ Очистка тестовых данных\n\n"
        "/bb_clear — удалить сделки, расходы, доходы и долги, но оставить партнёров\n"
        "/bb_delete_deal 1 — удалить одну конкретную сделку\n"
        "/bb_reset — удалить вообще всё, включая партнёров\n\n"
        "Команды очистки доступны только админу."
    )


@dp.message(Command("bb_join"))
async def join_handler(message: Message):
    data = user_data(message)

    add_or_update_partner(
        telegram_id=data["telegram_id"],
        username=data["username"],
        first_name=data["first_name"],
        last_name=data["last_name"]
    )

    username = f"@{data['username']}" if data["username"] else "не указан"

    await message.answer(
        "✅ Партнёр зарегистрирован\n\n"
        f"👤 Имя: {data['first_name'] or 'не указано'}\n"
        f"🔗 Username: {username}\n\n"
        "Теперь ты можешь добавлять сделки, расходы и доходы."
    )


@dp.message(Command("bb_me"))
async def me_handler(message: Message):
    partner = require_partner(message)

    if partner is None:
        await message.answer("⚠️ Ты ещё не зарегистрирован. Напиши /bb_join")
        return

    _, telegram_id, username, first_name, last_name, display_name, is_admin = partner

    await message.answer(
        "👤 Твой профиль\n\n"
        f"ID: {telegram_id}\n"
        f"Имя: {display_name}\n"
        f"Username: @{username if username else 'не указан'}\n"
        f"Роль: {'админ' if is_admin else 'партнёр'}"
    )


@dp.message(Command("bb_partners"))
async def partners_handler(message: Message):
    partners = get_partners()

    if not partners:
        await message.answer("👥 Партнёров пока нет. Каждый должен написать /bb_join")
        return

    text = "👥 Партнёры команды\n\n"

    for partner_id, telegram_id, username, display_name, is_admin in partners:
        role = "👑 админ" if is_admin else "🤝 партнёр"
        text += f"{partner_id}. {user_view(display_name, username)} — {role}\n"

    await message.answer(text)


@dp.message(Command("bb_new_deal"))
async def new_deal_handler(message: Message):
    partner = require_partner(message)

    if partner is None:
        await message.answer("⚠️ Сначала зарегистрируйся: /bb_join")
        return

    parts = message.text.split(maxsplit=1)

    if len(parts) < 2:
        await message.answer("Пример:\n/bb_new_deal iPhone 12 128GB Black")
        return

    title = parts[1].strip()
    deal_id = create_deal(title, message.from_user.id)

    await message.answer(
        "📦 Сделка создана\n\n"
        f"Номер: #{deal_id}\n"
        f"Название: {title}\n"
        "Статус: 🛒 куплен\n\n"
        "Теперь можно добавить расход:\n"
        f"/bb_expense {deal_id} 28000 покупка"
    )


@dp.message(Command("bb_deals"))
async def deals_handler(message: Message):
    deals = get_deals()

    if not deals:
        await message.answer("📦 Сделок пока нет.\n\nСоздать сделку: /bb_new_deal iPhone 12")
        return

    text = "📦 Последние сделки\n\n"

    for deal_id, title, status, is_closed, created_at in deals:
        closed_text = "закрыта" if is_closed else "открыта"
        text += f"#{deal_id} {deal_status_icon(status)} {title}\n"
        text += f"Статус: {status}, {closed_text}\n\n"

    await message.answer(text)


@dp.message(Command("bb_deal"))
async def deal_handler(message: Message):
    parts = message.text.split(maxsplit=1)

    if len(parts) < 2:
        await message.answer("Пример:\n/bb_deal 1")
        return

    try:
        deal_id = int(parts[1])
    except ValueError:
        await message.answer("Номер сделки должен быть числом.")
        return

    deal = get_deal(deal_id)

    if deal is None:
        await message.answer("⚠️ Сделка не найдена.")
        return

    deal_id, title, status, is_closed, created_at, closed_at, creator_name, creator_username = deal
    totals = get_deal_totals(deal_id)
    transactions = get_deal_transactions(deal_id)

    text = (
        f"📦 Сделка #{deal_id}\n\n"
        f"Название: {title}\n"
        f"Статус: {deal_status_icon(status)} {status}\n"
        f"Состояние: {'закрыта' if is_closed else 'открыта'}\n"
        f"Создал: {user_view(creator_name, creator_username)}\n\n"
        "💰 Финансы:\n"
        f"Расходы: {money(totals['expense'])} ₽\n"
        f"Доходы: {money(totals['income'])} ₽\n"
        f"Прибыль: {money(totals['profit'])} ₽\n\n"
    )

    if transactions:
        text += "🧾 Операции:\n"

        for transaction_id, transaction_type, amount, comment, created_at, display_name, username in transactions:
            text += (
                f"{type_icon(transaction_type)} — {money(amount)} ₽\n"
                f"Партнёр: {user_view(display_name, username)}\n"
                f"Комментарий: {comment or '-'}\n\n"
            )
    else:
        text += "🧾 Операций пока нет.\n"

    await message.answer(text)


@dp.message(Command("bb_status"))
async def status_handler(message: Message):
    parts = message.text.split(maxsplit=2)

    if len(parts) < 3:
        await message.answer(
            "Пример:\n/bb_status 1 ремонт\n\n"
            "Статусы: куплен, проверка, ремонт, готов, выставлен, бронь, продан, закрыт"
        )
        return

    try:
        deal_id = int(parts[1])
    except ValueError:
        await message.answer("Номер сделки должен быть числом.")
        return

    raw_status = parts[2].lower().strip()

    if raw_status not in ALLOWED_STATUSES:
        await message.answer(
            "⚠️ Неизвестный статус.\n\n"
            "Доступные статусы:\n"
            "куплен, проверка, ремонт, готов, выставлен, бронь, продан, закрыт"
        )
        return

    status = ALLOWED_STATUSES[raw_status]

    try:
        update_deal_status(deal_id, status)
        await message.answer(f"✅ Статус сделки #{deal_id} изменён на: {deal_status_icon(status)} {status}")
    except ValueError as error:
        await message.answer(f"⚠️ {error}")


async def deal_transaction_handler(message: Message, transaction_type: str):
    partner = require_partner(message)

    if partner is None:
        await message.answer("⚠️ Сначала зарегистрируйся: /bb_join")
        return

    parts = message.text.split(maxsplit=3)

    if len(parts) < 3:
        await message.answer(
            "Формат:\n"
            "/bb_expense 1 28000 покупка\n"
            "/bb_income 1 38000 продажа"
        )
        return

    try:
        deal_id = int(parts[1])
        amount = int(parts[2])
    except ValueError:
        await message.answer("Номер сделки и сумма должны быть числами.")
        return

    if amount <= 0:
        await message.answer("Сумма должна быть больше нуля.")
        return

    comment = parts[3] if len(parts) > 3 else ""

    try:
        add_deal_transaction(
            deal_id=deal_id,
            telegram_id=message.from_user.id,
            transaction_type=transaction_type,
            amount=amount,
            comment=comment
        )

        await message.answer(
            f"{type_icon(transaction_type)} добавлен\n\n"
            f"Сделка: #{deal_id}\n"
            f"Сумма: {money(amount)} ₽\n"
            f"Комментарий: {comment or '-'}"
        )

    except ValueError as error:
        await message.answer(f"⚠️ {error}")


@dp.message(Command("bb_expense"))
async def deal_expense_handler(message: Message):
    await deal_transaction_handler(message, "expense")


@dp.message(Command("bb_income"))
async def deal_income_handler(message: Message):
    await deal_transaction_handler(message, "income")


@dp.message(Command("bb_close"))
async def close_deal_handler(message: Message):
    parts = message.text.split(maxsplit=1)

    if len(parts) < 2:
        await message.answer("Пример:\n/bb_close 1")
        return

    try:
        deal_id = int(parts[1])
    except ValueError:
        await message.answer("Номер сделки должен быть числом.")
        return

    try:
        result = close_deal_and_calculate_payments(deal_id)

        text = (
            f"🏁 Сделка #{deal_id} закрыта\n\n"
            f"Расходы: {money(result['expense'])} ₽\n"
            f"Доходы: {money(result['income'])} ₽\n"
            f"Чистая прибыль: {money(result['profit'])} ₽\n"
            f"Доля каждого: {money(result['share'])} ₽\n\n"
        )

        if result["payments"]:
            text += "💳 Кто кому должен:\n"

            for payment in result["payments"]:
                text += f"• {payment['from']} → {payment['to']}: {money(payment['amount'])} ₽\n"
        else:
            text += "✅ Долгов по этой сделке нет.\n"

        text += "\nПосмотреть все долги: /bb_debts"

        await message.answer(text)

    except ValueError as error:
        await message.answer(f"⚠️ {error}")


@dp.message(Command("bb_debts"))
async def debts_handler(message: Message):
    payments = get_unpaid_payments()

    if not payments:
        await message.answer("✅ Активных долгов нет.")
        return

    text = "💳 Активные долги\n\n"

    for payment_id, deal_id, title, from_name, from_username, to_name, to_username, amount, created_at in payments:
        text += (
            f"#{payment_id} по сделке #{deal_id} — {title}\n"
            f"{user_view(from_name, from_username)} → {user_view(to_name, to_username)}\n"
            f"Сумма: {money(amount)} ₽\n"
            f"Отметить оплату: /bb_pay {payment_id}\n\n"
        )

    await message.answer(text)


@dp.message(Command("bb_pay"))
async def pay_handler(message: Message):
    parts = message.text.split(maxsplit=1)

    if len(parts) < 2:
        await message.answer("Пример:\n/bb_pay 1")
        return

    try:
        payment_id = int(parts[1])
    except ValueError:
        await message.answer("Номер долга должен быть числом.")
        return

    try:
        mark_payment_paid(payment_id)
        await message.answer(f"✅ Долг #{payment_id} отмечен как оплаченный.")
    except ValueError as error:
        await message.answer(f"⚠️ {error}")


@dp.message(Command("bb_stock"))
async def stock_handler(message: Message):
    deals = get_deals(only_open=True)

    if not deals:
        await message.answer("📦 Открытых сделок нет.")
        return

    text = "📦 Техника в работе\n\n"

    for deal_id, title, status, is_closed, created_at in deals:
        totals = get_deal_totals(deal_id)

        text += (
            f"#{deal_id} {deal_status_icon(status)} {title}\n"
            f"Статус: {status}\n"
            f"Вложено: {money(totals['expense'])} ₽\n"
            f"Получено: {money(totals['income'])} ₽\n\n"
        )

    await message.answer(text)


@dp.message(Command("bb_capital"))
async def capital_handler(message: Message):
    data = get_global_capital()

    await message.answer(
        "📊 Капитал\n\n"
        "Открытые сделки:\n"
        f"Деньги в товаре: {money(data['open_frozen'])} ₽\n"
        f"Расходы по открытым: {money(data['open_expense'])} ₽\n"
        f"Доходы по открытым: {money(data['open_income'])} ₽\n\n"
        "За всё время:\n"
        f"Всего расходов: {money(data['total_expense'])} ₽\n"
        f"Всего доходов: {money(data['total_income'])} ₽\n"
        f"Общая прибыль: {money(data['total_profit'])} ₽"
    )


async def report_handler(message: Message, days: int, title: str):
    data = get_report(days)

    await message.answer(
        f"📈 {title}\n\n"
        f"Создано сделок: {data['deals_created']}\n"
        f"Закрыто сделок: {data['deals_closed']}\n\n"
        f"Расходы: {money(data['expense'])} ₽\n"
        f"Доходы: {money(data['income'])} ₽\n"
        f"Прибыль периода: {money(data['profit'])} ₽\n\n"
        f"Кол-во расходов: {data['expense_count']}\n"
        f"Кол-во доходов: {data['income_count']}"
    )


@dp.message(Command("bb_day"))
async def day_report_handler(message: Message):
    await report_handler(message, 1, "Отчёт за день")


@dp.message(Command("bb_week"))
async def week_report_handler(message: Message):
    await report_handler(message, 7, "Отчёт за неделю")


@dp.message(Command("bb_month"))
async def month_report_handler(message: Message):
    await report_handler(message, 30, "Отчёт за месяц")


@dp.message(Command("bb_history"))
async def history_handler(message: Message):
    rows = get_history()

    if not rows:
        await message.answer("🧾 История пока пустая.")
        return

    text = "🧾 Последние операции\n\n"

    for created_at, deal_id, title, transaction_type, amount, comment, display_name, username in rows:
        text += (
            f"{type_icon(transaction_type)} — {money(amount)} ₽\n"
            f"Сделка #{deal_id}: {title}\n"
            f"Партнёр: {user_view(display_name, username)}\n"
            f"Комментарий: {comment or '-'}\n\n"
        )

    await message.answer(text)


@dp.message(Command("bb_clear"))
async def clear_test_data_handler(message: Message):
    admin = require_admin(message)

    if admin is None:
        await message.answer(
            "⛔ У тебя нет прав на очистку данных.\n\n"
            "Эта команда доступна только админу."
        )
        return

    clear_test_data()

    await message.answer(
        "🧹 Тестовые данные очищены\n\n"
        "Удалено:\n"
        "• все сделки;\n"
        "• все расходы и доходы;\n"
        "• все долги.\n\n"
        "Партнёры сохранены."
    )


@dp.message(Command("bb_reset"))
async def reset_all_data_handler(message: Message):
    admin = require_admin(message)

    if admin is None:
        await message.answer(
            "⛔ У тебя нет прав на полный сброс.\n\n"
            "Эта команда доступна только админу."
        )
        return

    reset_all_data()

    await message.answer(
        "⚠️ Полный сброс выполнен\n\n"
        "Удалено всё:\n"
        "• партнёры;\n"
        "• сделки;\n"
        "• расходы;\n"
        "• доходы;\n"
        "• долги.\n\n"
        "Теперь каждому партнёру нужно снова написать /bb_join."
    )


@dp.message(Command("bb_delete_deal"))
async def delete_deal_handler(message: Message):
    admin = require_admin(message)

    if admin is None:
        await message.answer(
            "⛔ У тебя нет прав на удаление сделок.\n\n"
            "Эта команда доступна только админу."
        )
        return

    parts = message.text.split(maxsplit=1)

    if len(parts) < 2:
        await message.answer("Формат команды:\n/bb_delete_deal 1")
        return

    try:
        deal_id = int(parts[1])
    except ValueError:
        await message.answer("Номер сделки должен быть числом.")
        return

    try:
        delete_deal_by_id(deal_id)

        await message.answer(
            f"🗑 Сделка #{deal_id} удалена\n\n"
            "Также удалены все расходы, доходы и долги по этой сделке."
        )

    except ValueError as error:
        await message.answer(f"⚠️ {error}")


@dp.message(Command("bb_export"))
async def export_handler(message: Message):
    data = get_export_data()

    wb = Workbook()

    ws = wb.active
    ws.title = "Сделки"
    ws.append(["ID", "Название", "Статус", "Закрыта", "Создана", "Закрыта дата"])

    for row in data["deals"]:
        ws.append(row)

    ws = wb.create_sheet("Операции")
    ws.append(["ID", "Сделка ID", "Сделка", "Партнёр", "Тип", "Сумма", "Комментарий", "Дата"])

    for row in data["transactions"]:
        ws.append(row)

    ws = wb.create_sheet("Долги")
    ws.append(["ID", "Сделка", "Кто должен", "Кому должен", "Сумма", "Оплачено", "Создан", "Оплачен"])

    for row in data["payments"]:
        ws.append(row)

    ws = wb.create_sheet("Партнёры")
    ws.append(["ID", "Telegram ID", "Username", "Имя", "Админ", "Создан"])

    for row in data["partners"]:
        ws.append(row)

    filename = f"resale_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    wb.save(filename)

    await message.answer_document(
        FSInputFile(filename),
        caption="📊 Excel-отчёт готов"
    )

    try:
        os.remove(filename)
    except OSError:
        pass


async def main():
    init_db()
    await dp.start_polling(bot)


if __name__ == "__main__":
    asyncio.run(main())
